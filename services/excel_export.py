from openpyxl import Workbook
from openpyxl.styles import PatternFill


def minutes_to_time(minutes):
    hours = minutes // 60
    minutes = minutes % 60
    return f"{hours:02}:{minutes:02}"


def export_excel(rows):

    wb = Workbook()
    ws = wb.active

    ws.title = "Attendance"

    ws.append([
        "تاریخ",
        "نوع روز",
        "ورود",
        "خروج",
        "ساعت حضور",
        "اضافه‌کار",
        "کسرکار"
    ])



    
    for r in rows:

        ws.append([
            r["work_date"],
            r["day_type"],
            r["check_in"],
            r["check_out"],
            minutes_to_time(r["work_minutes"]),
            minutes_to_time(r["overtime_minutes"]),
            minutes_to_time(r["shortage_minutes"])
        ])
 
        fills = {
            "مرخصی استحقاقی": PatternFill(
                fill_type="solid",
                start_color="C6EFCE",
                end_color="C6EFCE"
            ),

            "مرخصی ساعتی": PatternFill(
                fill_type="solid",
                start_color="C6EFCE",
                end_color="C6EFCE"
            ),

            "مرخصی استعلاجی": PatternFill(
                fill_type="solid",
                start_color="C6EFCE",
                end_color="C6EFCE"
            ),

            "مأموریت": PatternFill(
                fill_type="solid",
                start_color="D9EAF7",
                end_color="D9EAF7"
            ),

            "تعطیل رسمی": PatternFill(
                fill_type="solid",
                start_color="E7E6E6",
                end_color="E7E6E6"
            ),

            "غیبت": PatternFill(
                fill_type="solid",
                start_color="F8CBAD",
                end_color="F8CBAD"
            )
        }

        if r["day_type"] in fills:

            row_num = ws.max_row

            for cell in ws[row_num]:
                cell.fill = fills[r["day_type"]]
        
        
        
    wb.save("reports/attendance.xlsx")